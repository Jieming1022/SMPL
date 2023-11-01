import os
import pandas as pd
from tbparse import SummaryReader
from openpyxl import Workbook, load_workbook
from openpyxl.styles import DEFAULT_FONT
DEFAULT_FONT.name = 'Times New Roman'


def save(max_epoch, tensorboard_directory, file_name):
    reader = SummaryReader(tensorboard_directory)
    df_data = reader.scalars
    tags = reader.tags
    wb = Workbook()
    for index, tag in enumerate(tags['scalars']):
        tag = tag.replace('/', '_')
        if index == 0:
            wb.active.title = tag
        else:
            wb.create_sheet(tag)

    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    writer.book = wb
    writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

    for index, tag in enumerate(tags['scalars']):
        data = df_data[df_data['tag'] == tag][['step', 'value']]
        data.sort_values(by=['step'], inplace=True)
        if len(data) > max_epoch:
            data['epoch'] = (data['step'] / int(len(data) / max_epoch)).astype('int') + 1
            cols = data.columns.tolist()
            cols = cols[-1:] + cols[:-1]
            data = data[cols]
        else:
            data = data.rename({'step': 'epoch'}, axis=1)
        if '/' in tag:
            tag = tag.split('/')[1]
        else:
            tag = tag.replace('/', '_')
        data.to_excel(writer, sheet_name=tag, index=False)
    writer.save()